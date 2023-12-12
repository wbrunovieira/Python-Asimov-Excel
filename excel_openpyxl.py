from openpyxl import load_workbook

wb = load_workbook("teste.xlsx")
# Obter os nomes das planilhas
sheet_names = wb.sheetnames
sheet = wb.active
print(sheet_names)

# Acessar uma planilha específica
sheet = wb["Sheet1"]

# Obter o valor de uma célula específica
cell_value = sheet["A3"].value
print(cell_value)

sheet.cell(row=1, column=1).value = "Teste"

maxcolu =sheet.max_column
maxrow = sheet.max_row
print(maxcolu, maxrow)

for i in range(1, maxcolu+1):
    for j in range(1, maxrow+1):
        print(sheet.cell(row=j, column=i).value)
        
sheet.cell(row=1, column=2).value = 75

sheet.insert_rows(1)

wb.save("teste.xlsx")